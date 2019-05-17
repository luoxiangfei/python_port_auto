# -*- coding: utf-8 -*-
# @Author: 罗湘飞
# @Date  : 2019/2/23/023

import pyexcel_xls
import xlsxwriter
import openpyxl
from common.LogInfo import logger
class Excel(object):
    def __init__(self):
        self.log=logger
    def get_xls(self,url,sheet):
        """读取ecxel文件"""
        self.data_xls=pyexcel_xls.get_data(url)
        #这里读取excel文件，
        for get_sheet in self.data_xls.keys():
            if str(get_sheet)==sheet:
                return self.data_xls[get_sheet]
    def write_xls(self,listNum,xlsname="测试结果.xlsx",sheet="sheet"):
        try:
            self.workbook = xlsxwriter.Workbook (xlsname)  #创建一个excel文件
            self.worksheet = self.workbook.add_worksheet (sheet)    #在文件中创建一个名为TEST的sheet,不加名字默认为sheet1
            for row_num, row_data in enumerate (listNum):   #循环读取excel文件
                self.worksheet.write_row (row_num, 0, row_data)  #把每一行再写入到excel中
            self.workbook.close ()
            self.log.info("用例执行并保存成功")
        except Exception as e:
            self.log.info("用例保存失败,失败原因：{}".format(e))
    def write_newxls(self,data,xlsxname,sheet="测试结果"):
        '''生成新的xls文件'''
        try:
            workbook = openpyxl.Workbook (xlsxname)
            ws1 = workbook.create_sheet (sheet)
            for i in range (len (data)):
                ws1.append (data[i])
            workbook.save (filename=xlsxname)
            self.log.info("用例执行并保存为{}成功".format(xlsxname))
        except Exception as e:
            self.log.error("用例保存失败,失败原因：{}".format(e))
    def amend_xls(self,data,xlsxname,sheet="测试结果"):
        '''修改excel文件---增加sheet'''
        try:
            workbook = openpyxl.load_workbook (xlsxname)
            ws1 = workbook.create_sheet (sheet)  # 创建一个sheet
            for i in range (len (data)):
                ws1.append (data[i])
            workbook.save (filename=xlsxname)
            self.log.info("用例执行并保存为{}成功,sheet为{}".format(xlsxname,sheet))
        except Exception as e:
            self.log.error("用例保存失败,失败原因：{}".format(e))
if __name__ == '__main__':
    Excel().write_newxls([[1,2,3,4],[12,32,45,23],[11,22,33,44]],"luo.xlsx")